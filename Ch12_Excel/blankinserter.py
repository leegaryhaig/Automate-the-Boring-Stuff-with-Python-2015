import openpyxl, sys

# python blankRowInserter.py start_line blank_lines myProduce.xlsx
# First argument [3] is after which line to insert and second argument [2] is how many lines to insert, last is the excel sheet3

def blanklines(start_line, blank_lines, xlfile):

    wb = openpyxl.load_workbook(xlfile)
    sheet = wb.active
    tmpsheet = wb.create_sheet('tempsheet')

    # if not start_line.isnumeric():
    #     raise TypeError('Invalid data type for Start_line, please use an integer')
    #
    # if not blank_lines.isnumeric():
    #     raise TypeError('Invalid data type for Blank_line, please use an integer')

    start_line = int(start_line)
    blank_lines = int(blank_lines)
    new_start = start_line + blank_lines

    for row_num in range(1, start_line):
        for column_num in range(1, sheet.max_column + 1):
            tmpsheet.cell(row=row_num, column=column_num).value = sheet.cell(row=row_num, column=column_num).value

    for row_num in range(new_start, sheet.max_row + blank_lines + 1):
        for column_num in range(1, sheet.max_column + 1):
            tmpsheet.cell(row=row_num, column=column_num).value = sheet.cell(row=(row_num - blank_lines), column=column_num).value

    wb.save('insertLine.xlsx')
    print('Lines inserted, insertLine.xlsx updated')


try:
    blanklines(sys.argv[1], sys.argv[2], sys.argv[3])
except TypeError as err:
    raise ValueError('Invalid Argument Value') from err
except FileNotFoundError as e:
    print(e)
