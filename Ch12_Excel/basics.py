import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.create_sheet('sheet')

def openxldoc():
    print(type(wb))

def getsheets():
    print(wb.sheetnames)
    sheet = wb['Sheet3']
    print(sheet)
    print(sheet.title)
    anothersheet = wb.active
    print(anothersheet)

def getcells():
    print("command: sheet1['A1'] = " + str(sheet['A1']))
    print("command: sheet1['A1'].value = " + str(sheet['A1'].value))
    print('\n')
    c = sheet['B1']
    print("command: type(c) = " + str(type(c)))
    print("command: c.value = " + str(c.value))
    print('\n')
    print('Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value)
    print('Cell ' + c.coordinate + ' is ' + c.value)
    print("command: sheet1['C1'] = " + str(sheet['C1'].value))

def lettertonum():
    from openpyxl.utils import get_column_letter
    letter = get_column_letter(2)
    print(letter)

lettertonum()


def changefont():
    from openpyxl.styles import Font
    font_obj = Font(bold=True, italic=True)
    sheet['A1'].font = font_obj


def formulas():
    sheet['A1'] = '200'
    sheet['A2'] = '300'
    sheet['A3'] = '=SUM(A1:A2)'
    wb.save('writeformula.xlsx')

def adjustrowcolumn():
    sheet['A1'] = 'Tall row'
    sheet['B2'] = 'Wide column'
    sheet.row_dimensions[1] = '70'  # Set row 1 as 70 'points' (0-409max)
    sheet.column_dimensions['B'] = '20'  # Set column 'B' as 20 points (0-255max)
    wb.save('adjustcolumns.xlsx')

def mergecells():
    sheet.merge_cells('A1:D3')
    sheet['A1'] = 'Twelve Cells Merged together'  # To set values for merged cells - use the top-left cell
    sheet.merge_cells('C5:D5')
    sheet['C5'] = 'Two Cells Merged together'
    sheet.unmerge_cells('C5:D5')

def freezepanes():
    # For spreadsheets that are too large to be displayed at once its helpful to 'freeze' a few top rows or left columns
    sheet.freeze_panes = 'A2'  # Row 1 is frozen
    sheet.freeze_panes = 'B2'  # Column A is frozen
    sheet.freeze_panes = 'C1'  # Columns A and B
    sheet.freeze_panes = 'C2'  # Row 1 and Columns A and B

    sheet.freeze_panes = 'A1'  # No frozen panes
    sheet.freeze_panes = None  # No frozen panes







