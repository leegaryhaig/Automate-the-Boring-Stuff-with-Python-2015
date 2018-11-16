#! Python 3
import openpyxl
from openpyxl.utils import get_column_letter

# Write a program to read in the contents of several text files (you can make the text files yourself) and insert
# those contents into a spreadsheet, with one line of text per row. The lines of the first text file will be in the
# cells of column A, the lines of the second text file will be in the cells of column B, and so on.

file_start = 'lorem'
file_ext = '.txt'
wb = openpyxl.Workbook('texttosheet.xlsx')
sheet = wb.active
tmpsheet = wb.create_sheet('tmpsheet')

# Loops through lorem1.txt to lorem3.txt
for x in range(1, 4):
    filename = file_start + str(x) + file_ext
    print(filename)

    # Places the file in a file Object
    file_object = open(filename, 'r')
    print(type(file_object))

    # Copies the string of each line into a list stored in variable file_data
    file_list = file_object.readlines()
    print(file_list)

    # Loops through the list
    for i in range(1, len(file_list) + 1):
        tmpsheet.cell(row=i, column=get_column_letter(x)).value = file_list[i]



