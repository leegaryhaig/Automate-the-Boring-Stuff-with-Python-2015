import openpyxl, sys
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb.active
font_obj = Font(bold=True)


sheet.ma
def createTable():
    if len(sys.argv) != 2:
        raise Exception('You must provide the correct number of arguments')


    # Create the Keys, B1 - "sys.argv[x]" + 1, A2 - A + "sys.argv[x]"
    for key_length in range(1, (int(sys.argv[1]))+1):
        sheet[get_column_letter(key_length) + str(1)].value = key_length  # Row
        sheet[get_column_letter(key_length) + str(1)].font = font_obj

        sheet['A' + str(key_length)].value = key_length  # Column
        sheet['A' + str(key_length)].font = font_obj

    for column in range(2, (int(sys.argv[1]))+1):
        x = sheet[get_column_letter(column) + str(1)]

        for row in range(2, int(sys.argv[1])+1):
            y = sheet['A' + str(row)]
            answer = sheet[get_column_letter(column) + str(row)]

            answer.value = x.value * y.value

    wb.save('MultiplicationTable.xlsx')

try:
    createTable()

except Exception as err:
    print('An Exception happened: ' + str(err))


