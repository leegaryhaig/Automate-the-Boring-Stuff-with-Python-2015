#! Python 3
import openpyxl
from openpyxl.utils import get_column_letter

# This program will invert the row and columns ex: value at row 5 col 3 will be col 5, row 3

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
tmpsheet = wb.create_sheet('tempsheet')

# Loops through each row and column
for row_num in range(1, sheet.max_row):
    for column_num in range(1, sheet.max_column + 1):
        tmpsheet.cell(row=column_num, column=row_num).value = sheet.cell(row=row_num, column=column_num).value

wb.save('invertedsheet.xlsx')

