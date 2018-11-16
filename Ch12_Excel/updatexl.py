#! Python 3.
import openpyxl
from openpyxl.styles import Font


PRICE_UPDATES = {'celery': 1.19,
                 'garlic': 3.07,
                 'lemon': 1.27}

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
font_obj1 = Font(bold=True, italic=True)
formula_obj = '=SUM(' + 'D2:D' + str(sheet.max_row) + ')'

for row in range(2, sheet.max_row):  # Skip the first line and loops through all of the rows and columns
    produceName = sheet['A' + str(row)].value
    if produceName.lower() in PRICE_UPDATES:
        sheet['B' + str(row)].value = PRICE_UPDATES[produceName.lower()]
        sheet['B' + str(row)].font = font_obj1  # Applies font style

sheet['E2'] = formula_obj

wb.save('updatedProduceSales.xlsx')
