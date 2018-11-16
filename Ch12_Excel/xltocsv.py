#! Python3
import openpyxl, csv, os

csvRow = []

# TODO: Create directory to store the csv files
os.makedirs('csvFile', exist_ok=True)


# TODO: Loop through the current directory for excel files
for excelFile in os.listdir('.'):
    if excelFile.endswith('.xlsx'):
        # Opens the Excel file and last active sheet
        wb = openpyxl.load_workbook(excelFile)
        sheet = wb.active

        # Creates csv equivalent file inside of 'cvsFile' directory
        csvName = excelFile.rstrip('.xlsx') + str('.csv')
        csvFileObj = open('csvFile/' + csvName, 'w', newline='')

        for row_num in range(1, sheet.max_row):
            for column_num in range(1, sheet.max_column + 1):
                cellObj = sheet.cell(row=row_num, column=column_num).value
                csvRow.append(cellObj)
            csvWriter = csv.writer(csvFileObj)
            csvWriter.writerow(csvRow)
            csvRow = []

        wb.close()
        csvFileObj.close()


